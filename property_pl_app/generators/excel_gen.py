"""
Parameterized Excel generator for Property P&L + Portfolio Dashboard.
Supports 1–10 properties, any FY start month, any FY range.
"""

import io
import calendar
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle

# ── Universal constants (not theme-dependent) ──────────────────────────────────
INPUT_BLUE = "FF0070C0"
GREEN_LINK = "FF00B050"
FY_YELLOW  = "FFFFC000"   # period marker – FY total column header
CY_YELLOW  = "FFFFE699"   # period marker – CY total column header
TMPL_GREY  = "FFF2F2F2"   # no-data template cell
WHITE      = "FFFFFFFF"
BLACK      = "FF000000"

NUM_DOLLAR  = '$#,##0.00;($#,##0.00);"-"'
NUM_INT     = '$#,##0;($#,##0);"-"'
NUM_PERCENT = '0.0%;(0.0%);"-"'
NUM_DSCR    = '0.00x;(0.00x);"-"'

# ── Themes ─────────────────────────────────────────────────────────────────────
# Each theme defines structural chrome AND semantic row colors.
# Semantic keys:
#   income_bg/fg        – item rows in INCOME section
#   income_tot_bg/fg    – Total Income row
#   expense_bg/fg       – item rows in OPEX / UTILITIES / FINANCING
#   expense_tot_bg/fg   – Total Operating Expenses / Utilities / Financing rows
#   net_bg/fg           – NOI, NET PROFIT, Net Cash Flow, NOI Margin %
#   cashflow_bg/fg      – Cash Flow item rows (EFT received, payments)
#   kpi_row_bg          – KPI table metric-label cells
THEMES = {
    'navy': {
        # Chrome
        'header':          'FF1F3864',
        'section':         'FF2F5496',
        'header_text':     'FFFFFFFF',
        'section_text':    'FFFFFFFF',
        # Semantic
        'income_bg':       'FFEBF5EB', 'income_fg':       'FF1B5E20',
        'income_tot_bg':   'FFC8E6C9', 'income_tot_fg':   'FF1B5E20',
        'expense_bg':      'FFFEECEC', 'expense_fg':      'FFB71C1C',
        'expense_tot_bg':  'FFFFCDD2', 'expense_tot_fg':  'FFB71C1C',
        'net_bg':          'FFEBF3FB', 'net_fg':          'FF1A237E',
        'cashflow_bg':     'FFECE8F5', 'cashflow_fg':     'FF4A148C',
        'kpi_row_bg':      'FFDAE3F3',
    },
    'sage': {
        # Chrome
        'header':          'FF2E4057',
        'section':         'FF445566',
        'header_text':     'FFFFFFFF',
        'section_text':    'FFFFFFFF',
        # Semantic
        'income_bg':       'FFF0FAF4', 'income_fg':       'FF2D6A4F',
        'income_tot_bg':   'FFB7E4C7', 'income_tot_fg':   'FF2D6A4F',
        'expense_bg':      'FFFDF0F0', 'expense_fg':      'FFA93226',
        'expense_tot_bg':  'FFFFC8C8', 'expense_tot_fg':  'FFA93226',
        'net_bg':          'FFE8F4F8', 'net_fg':          'FF1A6B8A',
        'cashflow_bg':     'FFFEF9E7', 'cashflow_fg':     'FFB7770D',
        'kpi_row_bg':      'FFD6EAF8',
    },
    'charcoal': {
        # Chrome
        'header':          'FF2C2C2C',
        'section':         'FF404040',
        'header_text':     'FFFFFFFF',
        'section_text':    'FFFF9F0A',   # amber text on dark section
        # Semantic
        'income_bg':       'FFF2FCF4', 'income_fg':       'FF1A6B2A',
        'income_tot_bg':   'FFB9F0C5', 'income_tot_fg':   'FF1A6B2A',
        'expense_bg':      'FFFDF2F2', 'expense_fg':      'FFC0392B',
        'expense_tot_bg':  'FFFFC8C0', 'expense_tot_fg':  'FFC0392B',
        'net_bg':          'FFF0F4FA', 'net_fg':          'FF1565C0',
        'cashflow_bg':     'FFFAF5E4', 'cashflow_fg':     'FFB05000',
        'kpi_row_bg':      'FFD6E4F7',
    },
}

# ── Semantic row classification ────────────────────────────────────────────────
SEMANTIC_TYPE: dict[str, str] = {
    # Income
    'Rental Income': 'income', 'Other Income': 'income', 'Excess Bill Shares': 'income',
    'Total Income': 'income_tot',
    # Operating expenses
    'Management Fees': 'expense', 'Letting Fees': 'expense', 'Council Rates': 'expense',
    'Land Tax': 'expense', 'Strata / Body Corporate': 'expense',
    'Building Insurance': 'expense', 'Maintenance & Repairs': 'expense',
    'Cleaning': 'expense', 'Advertising': 'expense', 'Miscellaneous': 'expense',
    'Total Operating Expenses': 'expense_tot',
    # Utilities
    'Electricity': 'expense', 'Water': 'expense', 'Gas': 'expense', 'Internet': 'expense',
    'Total Utilities': 'expense_tot',
    # Financing
    'Mortgage Interest': 'expense',
    'Total Financing Cost': 'expense_tot',
    # Net / KPI
    'NOI (Net Operating Income)': 'net', 'NOI Margin %': 'net',
    'NET PROFIT / (LOSS)': 'net',
    # Cash flow
    'Cash Received (EFT)': 'cashflow', 'Less: Utilities Paid': 'cashflow',
    'Less: Mortgage Repayment': 'cashflow', 'Net Cash Flow': 'net',
    'Principal Repaid': 'cashflow',
}


def _sem(label: str, th: dict) -> tuple:
    """Return (Fill, fg_hex) for a label based on semantic type and active theme."""
    s = SEMANTIC_TYPE.get(label, 'neutral')
    if s == 'income':      return Fill(th['income_bg']),      th['income_fg']
    if s == 'income_tot':  return Fill(th['income_tot_bg']),  th['income_tot_fg']
    if s == 'expense':     return Fill(th['expense_bg']),     th['expense_fg']
    if s == 'expense_tot': return Fill(th['expense_tot_bg']), th['expense_tot_fg']
    if s == 'net':         return Fill(th['net_bg']),         th['net_fg']
    if s == 'cashflow':    return Fill(th['cashflow_bg']),    th['cashflow_fg']
    return None, BLACK


# ── P&L structure ─────────────────────────────────────────────────────────────
PL_STRUCTURE = [
    ('section', 'INCOME'),
    ('item',    'Rental Income'),
    ('item',    'Other Income'),
    ('item',    'Excess Bill Shares'),
    ('total',   'Total Income'),
    ('blank',   None),
    ('section', 'OPERATING EXPENSES'),
    ('item',    'Management Fees'),
    ('item',    'Letting Fees'),
    ('item',    'Council Rates'),
    ('item',    'Land Tax'),
    ('item',    'Strata / Body Corporate'),
    ('item',    'Building Insurance'),
    ('item',    'Maintenance & Repairs'),
    ('item',    'Cleaning'),
    ('item',    'Advertising'),
    ('item',    'Miscellaneous'),
    ('total',   'Total Operating Expenses'),
    ('blank',   None),
    ('kpi',     'NOI (Net Operating Income)'),
    ('kpi',     'NOI Margin %'),
    ('blank',   None),
    ('section', 'UTILITIES'),
    ('item',    'Electricity'),
    ('item',    'Water'),
    ('item',    'Gas'),
    ('item',    'Internet'),
    ('total',   'Total Utilities'),
    ('blank',   None),
    ('section', 'FINANCING'),
    ('item',    'Mortgage Interest'),
    ('total',   'Total Financing Cost'),
    ('blank',   None),
    ('kpi',     'NET PROFIT / (LOSS)'),
    ('blank',   None),
    ('section', 'CASH FLOW'),
    ('item',    'Cash Received (EFT)'),
    ('item',    'Less: Utilities Paid'),
    ('item',    'Less: Mortgage Repayment'),
    ('total',   'Net Cash Flow'),
    ('item',    'Principal Repaid'),
]

# Key row labels for formula construction
LABEL_ROW: dict[str, int] = {}   # filled during build

INCOME_ITEMS    = ['Rental Income', 'Other Income', 'Excess Bill Shares']
OPEX_ITEMS      = ['Management Fees', 'Letting Fees', 'Council Rates', 'Land Tax',
                   'Strata / Body Corporate', 'Building Insurance',
                   'Maintenance & Repairs', 'Cleaning', 'Advertising', 'Miscellaneous']
UTILITY_ITEMS   = ['Electricity', 'Water', 'Gas', 'Internet']
FINANCING_ITEMS = ['Mortgage Interest']
CF_ITEMS        = ['Cash Received (EFT)', 'Less: Utilities Paid',
                   'Less: Mortgage Repayment']


def _derive_cy_labels(fy_labels: list[str]) -> list[int]:
    years: set[int] = set()
    for fy in fy_labels:
        base = int(fy.split('-')[0])
        years.add(base)
        years.add(base + 1)
    return sorted(years, reverse=True)


def _build_pl_rows() -> list[tuple[str, str | None]]:
    rows = []
    row_idx = 5
    for rtype, label in PL_STRUCTURE:
        if rtype == 'blank':
            rows.append(('blank', None))
            row_idx += 1
        else:
            rows.append((rtype, label))
            LABEL_ROW[label] = row_idx
            row_idx += 1
    return rows, row_idx


def _fy_months(fy_start_month: int) -> list[int]:
    end_month = (fy_start_month - 2) % 12 + 1
    months = []
    m = end_month
    for _ in range(12):
        months.append(m)
        m = (m - 2) % 12 + 1
    return months


def _month_label(month: int, fy_label: str, fy_start: int) -> str:
    base_year = int(fy_label.split('-')[0])
    end_year  = int(fy_label.split('-')[1]) if len(fy_label) == 7 else base_year + 1
    mo_abbr   = calendar.month_abbr[month]
    year      = base_year if month >= fy_start else end_year
    return f'{mo_abbr}-{str(year)[2:]}'


# ── Cell helpers ───────────────────────────────────────────────────────────────
def F(bold=False, color=BLACK, size=9, italic=False):
    return Font(name='Arial', bold=bold, color=color, size=size, italic=italic)

def Fill(c):
    return PatternFill("solid", fgColor=c)

def Aln(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def Border1(color='FFB8CCE4'):
    s = Side(style='thin', color=color)
    return Border(bottom=s)

def col(n):
    return get_column_letter(n)

def wcell(ws, r, c, val, font=None, fill=None, aln=None, num_fmt=None):
    cell = ws.cell(r, c, val)
    if font:    cell.font = font
    if fill:    cell.fill = fill
    if aln:     cell.alignment = aln
    if num_fmt: cell.number_format = num_fmt
    return cell

def mcell(ws, r, c1, c2, val, font=None, fill=None, aln=None):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(r, c1, val)
    if font: cell.font = font
    if fill: cell.fill = fill
    if aln:  cell.alignment = aln
    return cell


# ──────────────────────────────────────────────────────────────────────────────
# MAIN ENTRY: build_workbook
# ──────────────────────────────────────────────────────────────────────────────
def build_workbook(
    properties: list[dict],
    fy_start_month: int = 7,
    fy_labels: list[str] | None = None,
    purchase_info: dict | None = None,
    theme: str = 'navy',          # 'navy' | 'sage' | 'charcoal'
) -> bytes:
    """Build and return xlsx bytes."""

    global LABEL_ROW
    LABEL_ROW = {}

    th = THEMES.get(theme, THEMES['navy'])

    if fy_labels is None:
        fy_labels = ['2029-30','2028-29','2027-28','2026-27','2025-26','2024-25']

    purchase_info = purchase_info or {}
    wb = Workbook()
    wb.remove(wb.active)

    prop_tabs = []

    # ─────────────────────────────────────────────────────────────────────────
    # BUILD EACH PROPERTY TAB
    # ─────────────────────────────────────────────────────────────────────────
    for prop in properties:
        prop_name = prop.get('name', 'Property')
        tab_name  = prop.get('tab', prop_name[:31])
        prop_data = prop.get('data', {})

        ws = wb.create_sheet(tab_name)
        prop_tabs.append(tab_name)

        pl_rows, next_row = _build_pl_rows()

        # ── Build column layout ────────────────────────────────────────────
        fy_total_col   = {}
        fy_month_cols  = {}
        fy_month_range = {}

        month_seq = _fy_months(fy_start_month)

        c = 2
        for fy in fy_labels:
            fy_total_col[fy] = c
            c += 1
            fy_month_cols[fy] = {}
            start = c
            for mo in month_seq:
                fy_month_cols[fy][mo] = c
                c += 1
            fy_month_range[fy] = (start, c - 1)

        cy_labels  = _derive_cy_labels(fy_labels)
        cy_col     = {cy: c + i for i, cy in enumerate(cy_labels)}
        total_cols = c + len(cy_labels) - 1

        # period_to_col_map maps 'FY 2025-26' / 'CY 2025' → actual column index
        period_to_col_map = {}
        for fy in fy_labels:
            period_to_col_map[f'FY {fy}'] = fy_total_col[fy]
        for cy in cy_labels:
            period_to_col_map[f'CY {cy}'] = cy_col[cy]

        last_col = total_cols   # rightmost column with content

        # ── Row 1: Title ───────────────────────────────────────────────────
        ws.row_dimensions[1].height = 22
        mcell(ws, 1, 1, min(last_col, 20), prop_name,
              F(bold=True, color=th['header_text'], size=12),
              Fill(th['header']), Aln('center'))

        # ── Row 2: Legend ──────────────────────────────────────────────────
        ws.row_dimensions[2].height = 14
        legend = ('🟢 Income  🔴 Expenses  🔵 Net/Profit  🟣 Cash Flow  '
                  '│  Yellow = FY Total  │  Lt.Yellow = CY Total  │  Grey = Template')
        mcell(ws, 2, 1, min(last_col, 20), legend,
              F(size=8, italic=True, color='FF595959'), Fill('FFF8F8F8'), Aln('center'))

        # ── Row 3: spacer ──────────────────────────────────────────────────
        ws.row_dimensions[3].height = 5

        # ── Row 4: Column headers ──────────────────────────────────────────
        ws.row_dimensions[4].height = 32
        wcell(ws, 4, 1, 'Category',
              F(bold=True, color=th['header_text'], size=9),
              Fill(th['header']), Aln('center', wrap=True))

        for fy in fy_labels:
            is_tmpl = not any((y, m) in prop_data for m in range(1, 13)
                              for y in range(2020, 2040)
                              if f'{y}-{str(y+1)[2:]}' == fy or
                              f'{y-1}-{str(y)[2:]}' == fy)
            fy_hdr_bg = TMPL_GREY if is_tmpl else FY_YELLOW

            wcell(ws, 4, fy_total_col[fy], f'FY {fy}\nTotal',
                  F(bold=True, size=8), Fill(fy_hdr_bg), Aln('center', wrap=True))

            for mo in month_seq:
                lbl = _month_label(mo, fy, fy_start_month)
                base_yr  = int(fy.split('-')[0])
                yr       = base_yr if mo >= fy_start_month else base_yr + 1
                mo_bg    = TMPL_GREY if (yr, mo) not in prop_data else WHITE
                wcell(ws, 4, fy_month_cols[fy][mo], lbl,
                      F(bold=True, size=8), Fill(mo_bg), Aln('center', wrap=True))

        for cy in cy_labels:
            wcell(ws, 4, cy_col[cy], f'CY {cy}',
                  F(bold=True, size=8), Fill(CY_YELLOW), Aln('center', wrap=True))

        # ── Column outline grouping (monthly cols collapse per FY) ─────────
        for fy in fy_labels:
            s, e = fy_month_range[fy]
            for ci in range(s, e + 1):
                ws.column_dimensions[col(ci)].outlineLevel = 1
        ws.sheet_properties.outlinePr.summaryRight = False

        # ── Column widths ──────────────────────────────────────────────────
        ws.column_dimensions['A'].width = 28
        for fy in fy_labels:
            ws.column_dimensions[col(fy_total_col[fy])].width = 12
            for mo in month_seq:
                ws.column_dimensions[col(fy_month_cols[fy][mo])].width = 10
        for cy in cy_labels:
            ws.column_dimensions[col(cy_col[cy])].width = 12

        ws.freeze_panes = 'B5'

        # ── Data rows ──────────────────────────────────────────────────────
        for rtype, label in pl_rows:
            excel_row = LABEL_ROW.get(label)
            if rtype == 'blank' or excel_row is None:
                continue

            ws.row_dimensions[excel_row].height = 16

            sem_fill, sem_fg = _sem(label, th)

            # ── FIX 1: Section rows – fill every individual cell, no merge ─
            if rtype == 'section':
                sec_fill = Fill(th['section'])
                sec_font_label = F(bold=True, color=th['section_text'], size=9)
                sec_font_empty = F(bold=False, color=th['section_text'], size=9)
                for ci in range(1, last_col + 1):
                    wcell(ws, excel_row, ci,
                          label if ci == 1 else None,
                          font=sec_font_label if ci == 1 else sec_font_empty,
                          fill=sec_fill,
                          aln=Aln('left'))
                continue

            # Category label (col A) – semantic color
            if rtype in ('total', 'kpi'):
                lbl_font = F(bold=True, color=sem_fg or BLACK, size=9)
            else:
                lbl_font = F(size=9, color=sem_fg or BLACK)
            wcell(ws, excel_row, 1, label,
                  font=lbl_font, fill=sem_fill, aln=Aln('left'))

            # Check which FYs have any data (for FY total coloring)
            fy_has_data_map = {}
            for fy in fy_labels:
                base_yr = int(fy.split('-')[0])
                fy_has_data_map[fy] = any(
                    (base_yr if mo >= fy_start_month else base_yr + 1, mo) in prop_data
                    for mo in month_seq
                )

            # ── Monthly cells (white/grey – no semantic noise) ─────────────
            for fy in fy_labels:
                for mo in month_seq:
                    mc_idx  = fy_month_cols[fy][mo]
                    base_yr = int(fy.split('-')[0])
                    yr      = base_yr if mo >= fy_start_month else base_yr + 1
                    cell_val = prop_data.get((yr, mo), {}).get(label)
                    has_data = (yr, mo) in prop_data

                    if rtype == 'item':
                        if cell_val is not None:
                            val = cell_val
                            fc  = F(size=9, color=INPUT_BLUE)
                        else:
                            val = None
                            fc  = F(size=9, color='FF999999')
                    elif rtype == 'total':
                        item_list = (INCOME_ITEMS if label == 'Total Income' else
                                     OPEX_ITEMS if label == 'Total Operating Expenses' else
                                     UTILITY_ITEMS if label == 'Total Utilities' else
                                     FINANCING_ITEMS if label == 'Total Financing Cost' else
                                     CF_ITEMS if label == 'Net Cash Flow' else [])
                        rows_ref = ','.join(
                            f'{col(mc_idx)}{LABEL_ROW[i]}'
                            for i in item_list if i in LABEL_ROW
                        )
                        val = f'=SUM({rows_ref})' if rows_ref else None
                        fc  = F(size=9, color=sem_fg or BLACK)
                    elif rtype == 'kpi':
                        ti    = LABEL_ROW.get('Total Income')
                        tox   = LABEL_ROW.get('Total Operating Expenses')
                        noi_r = LABEL_ROW.get('NOI (Net Operating Income)')
                        tu_r  = LABEL_ROW.get('Total Utilities')
                        tf_r  = LABEL_ROW.get('Total Financing Cost')
                        cr    = col(mc_idx)
                        if label == 'NOI (Net Operating Income)':
                            val = f'={cr}{ti}-{cr}{tox}' if ti and tox else None
                        elif label == 'NOI Margin %':
                            val = f'=IFERROR({cr}{noi_r}/{cr}{ti},"-")' if noi_r and ti else None
                        elif label == 'NET PROFIT / (LOSS)':
                            val = (f'={cr}{noi_r}-{cr}{tu_r}-{cr}{tf_r}'
                                   if all([noi_r, tu_r, tf_r]) else None)
                        else:
                            val = None
                        fc = F(size=9, color=sem_fg or BLACK)
                    else:
                        val, fc = None, F(size=9)

                    # Monthly cells: white if data, grey if template
                    mo_bg = Fill(TMPL_GREY) if not has_data else Fill(WHITE)
                    nm = (NUM_PERCENT if label == 'NOI Margin %' else
                          NUM_INT    if rtype in ('item', 'total', 'kpi') else None)
                    wcell(ws, excel_row, mc_idx, val, font=fc, fill=mo_bg,
                          aln=Aln('right'), num_fmt=nm)

                # ── FY Total column – semantic if data, FY_YELLOW if template ─
                ft_col   = fy_total_col[fy]
                s_col, e_col = fy_month_range[fy]
                cr       = col(ft_col)
                has_fy   = fy_has_data_map[fy]
                ft_bg    = sem_fill if (has_fy and sem_fill) else Fill(FY_YELLOW)
                ft_fg    = sem_fg if (has_fy and sem_fg) else BLACK

                if rtype == 'item':
                    ft_val = f'=SUM({col(s_col)}{excel_row}:{col(e_col)}{excel_row})'
                elif rtype == 'total':
                    item_list = (INCOME_ITEMS if label == 'Total Income' else
                                 OPEX_ITEMS   if label == 'Total Operating Expenses' else
                                 UTILITY_ITEMS if label == 'Total Utilities' else
                                 FINANCING_ITEMS if label == 'Total Financing Cost' else
                                 CF_ITEMS if label == 'Net Cash Flow' else [])
                    rows_ref = ','.join(
                        f'{cr}{LABEL_ROW[i]}' for i in item_list if i in LABEL_ROW
                    )
                    ft_val = f'=SUM({rows_ref})' if rows_ref else None
                elif rtype == 'kpi':
                    ti    = LABEL_ROW.get('Total Income')
                    tox   = LABEL_ROW.get('Total Operating Expenses')
                    noi_r = LABEL_ROW.get('NOI (Net Operating Income)')
                    tu_r  = LABEL_ROW.get('Total Utilities')
                    tf_r  = LABEL_ROW.get('Total Financing Cost')
                    if label == 'NOI (Net Operating Income)':
                        ft_val = f'={cr}{ti}-{cr}{tox}' if ti and tox else None
                    elif label == 'NOI Margin %':
                        ft_val = f'=IFERROR({cr}{noi_r}/{cr}{ti},"-")' if noi_r and ti else None
                    elif label == 'NET PROFIT / (LOSS)':
                        ft_val = (f'={cr}{noi_r}-{cr}{tu_r}-{cr}{tf_r}'
                                  if all([noi_r, tu_r, tf_r]) else None)
                    else:
                        ft_val = None
                else:
                    ft_val = None

                nm = NUM_PERCENT if label == 'NOI Margin %' else NUM_INT
                wcell(ws, excel_row, ft_col, ft_val,
                      font=F(size=9, bold=True, color=ft_fg),
                      fill=ft_bg, aln=Aln('right'), num_fmt=nm)

            # ── CY Total columns – always semantic ─────────────────────────
            for cy in cy_labels:
                cy_refs = []
                for fy in fy_labels:
                    base_yr = int(fy.split('-')[0])
                    for mo in month_seq:
                        yr = base_yr if mo >= fy_start_month else base_yr + 1
                        if yr == cy:
                            cy_refs.append(f'{col(fy_month_cols[fy][mo])}{excel_row}')

                if rtype in ('item', 'total') and cy_refs:
                    cy_val = f'=SUM({",".join(cy_refs)})'
                elif rtype == 'kpi' and label == 'NOI (Net Operating Income)' and cy_refs:
                    ti_r  = LABEL_ROW.get('Total Income')
                    tox_r = LABEL_ROW.get('Total Operating Expenses')
                    cy_inc = [f'{col(fy_month_cols[fy][mo])}{ti_r}'
                              for fy in fy_labels for mo in month_seq
                              if ((int(fy.split('-')[0]) if mo >= fy_start_month
                                   else int(fy.split('-')[0]) + 1) == cy) if ti_r]
                    cy_opx = [f'{col(fy_month_cols[fy][mo])}{tox_r}'
                              for fy in fy_labels for mo in month_seq
                              if ((int(fy.split('-')[0]) if mo >= fy_start_month
                                   else int(fy.split('-')[0]) + 1) == cy) if tox_r]
                    cy_val = (f'=SUM({",".join(cy_inc)})-SUM({",".join(cy_opx)})'
                              if cy_inc and cy_opx else None)
                elif rtype == 'kpi' and label == 'NET PROFIT / (LOSS)' and cy_refs:
                    noi_r = LABEL_ROW.get('NOI (Net Operating Income)')
                    tu_r  = LABEL_ROW.get('Total Utilities')
                    tf_r  = LABEL_ROW.get('Total Financing Cost')
                    refs_noi = [f'{col(fy_month_cols[fy][mo])}{noi_r}'
                                for fy in fy_labels for mo in month_seq
                                if ((int(fy.split('-')[0]) if mo >= fy_start_month
                                     else int(fy.split('-')[0]) + 1) == cy) if noi_r]
                    refs_tu  = [f'{col(fy_month_cols[fy][mo])}{tu_r}'
                                for fy in fy_labels for mo in month_seq
                                if ((int(fy.split('-')[0]) if mo >= fy_start_month
                                     else int(fy.split('-')[0]) + 1) == cy) if tu_r]
                    refs_tf  = [f'{col(fy_month_cols[fy][mo])}{tf_r}'
                                for fy in fy_labels for mo in month_seq
                                if ((int(fy.split('-')[0]) if mo >= fy_start_month
                                     else int(fy.split('-')[0]) + 1) == cy) if tf_r]
                    cy_val = (f'=SUM({",".join(refs_noi)})-SUM({",".join(refs_tu)})-SUM({",".join(refs_tf)})'
                              if refs_noi else None)
                elif rtype == 'kpi' and label == 'NOI Margin %':
                    cy_val = None
                else:
                    cy_val = None

                cy_bg = sem_fill if sem_fill else Fill(CY_YELLOW)
                cy_fg = sem_fg if sem_fg else BLACK
                nm = NUM_PERCENT if label == 'NOI Margin %' else NUM_INT
                wcell(ws, excel_row, cy_col[cy], cy_val,
                      font=F(size=9, bold=(rtype in ('total', 'kpi')), color=cy_fg),
                      fill=cy_bg, aln=Aln('right'), num_fmt=nm)

        # ── TABLE A: KPI summary ───────────────────────────────────────────
        # FIX 2: KPI table columns placed at the same indices as P&L FY/CY total columns
        ta_start = next_row + 2
        kpi_periods = (
            [f'FY {fy}' for fy in fy_labels] +
            [f'CY {cy}' for cy in cy_labels]
        )

        r = ta_start

        # Banner row – fill every cell individually across full width
        ws.row_dimensions[r].height = 18
        banner_font  = F(bold=True, color=th['header_text'], size=10)
        banner_fill  = Fill(th['header'])
        banner_label = F(bold=True, color=th['header_text'], size=10)
        for ci in range(1, last_col + 1):
            wcell(ws, r, ci,
                  f'KEY PERFORMANCE INDICATORS — {prop_name}' if ci == 1 else None,
                  font=banner_label if ci == 1 else banner_font,
                  fill=banner_fill,
                  aln=Aln('left' if ci == 1 else 'center'))
        r += 1

        # Column-header row – 'Metric' at col 1, then each period at its actual column
        ws.row_dimensions[r].height = 30
        wcell(ws, r, 1, 'Metric',
              F(bold=True, color=th['header_text'], size=9),
              Fill(th['section']), Aln('center', wrap=True))
        for pk in kpi_periods:
            actual_c = period_to_col_map[pk]
            is_cy    = pk.startswith('CY')
            bg       = CY_YELLOW if is_cy else FY_YELLOW
            wcell(ws, r, actual_c, pk, F(bold=True, size=9), Fill(bg),
                  Aln('center', wrap=True))
        r += 1

        kpi_items = [
            ('Gross Rental Income', 'Total Income',               NUM_INT),
            ('NOI',                 'NOI (Net Operating Income)',  NUM_INT),
            ('NOI Margin %',        'NOI Margin %',               NUM_PERCENT),
            ('Financing Cost',      'Total Financing Cost',        NUM_INT),
            ('Net Profit / (Loss)', 'NET PROFIT / (LOSS)',         NUM_INT),
            ('Net Cash Flow',       'Net Cash Flow',               NUM_INT),
        ]

        for kpi_disp, pl_label, nm in kpi_items:
            ws.row_dimensions[r].height = 15
            pl_row   = LABEL_ROW.get(pl_label)
            row_sem_fill, row_sem_fg = _sem(pl_label, th)
            wcell(ws, r, 1, kpi_disp,
                  F(bold=True, size=9, color=row_sem_fg or BLACK),
                  row_sem_fill or Fill(th['kpi_row_bg']),
                  Aln('left'))
            for pk in kpi_periods:
                actual_c = period_to_col_map[pk]
                is_cy    = pk.startswith('CY')
                bg       = CY_YELLOW if is_cy else FY_YELLOW
                formula  = f'={col(actual_c)}{pl_row}' if pl_row else None
                wcell(ws, r, actual_c, formula, F(size=9), Fill(bg),
                      Aln('right'), nm)
            r += 1

        # DSCR row
        ws.row_dimensions[r].height = 15
        noi_r  = LABEL_ROW.get('NOI (Net Operating Income)')
        fin_r  = LABEL_ROW.get('Total Financing Cost')
        prin_r = LABEL_ROW.get('Principal Repaid')
        wcell(ws, r, 1, 'DSCR (NOI / Debt Service)',
              F(bold=True, size=9), Fill(th['kpi_row_bg']), Aln('left'))
        for pk in kpi_periods:
            actual_c = period_to_col_map[pk]
            is_cy    = pk.startswith('CY')
            bg       = CY_YELLOW if is_cy else FY_YELLOW
            cr       = col(actual_c)
            if noi_r and fin_r:
                formula = (f'=IFERROR({cr}{noi_r}/({cr}{fin_r}+{cr}{prin_r}),"-")'
                           if prin_r else
                           f'=IFERROR({cr}{noi_r}/{cr}{fin_r},"-")')
            else:
                formula = None
            wcell(ws, r, actual_c, formula, F(size=9), Fill(bg),
                  Aln('right'), NUM_DSCR)
        r += 1

    # ─────────────────────────────────────────────────────────────────────────
    # SUMMARY TAB
    # ─────────────────────────────────────────────────────────────────────────
    ws_s = wb.create_sheet('Summary')

    LABEL_ROW.clear()
    _build_pl_rows()

    cy_labels = _derive_cy_labels(fy_labels)
    all_fy    = ['FY ' + f for f in fy_labels]
    all_cy    = [f'CY {cy}' for cy in cy_labels]
    all_per   = all_fy + all_cy

    n_props = len(prop_tabs)

    TB_HEADERS = ['Property', 'Address', 'Purchase Price ($)',
                  'Purchase Date', 'Current Value ($)', 'Mortgage Balance ($)',
                  'Equity ($)', 'LVR (%)',
                  'Gross Yield (%)\n(FY Actual)',
                  'Net Yield (%)\n(FY Actual)',
                  'DSCR\n(FY Actual)']
    TB_LAST = len(TB_HEADERS)

    r = 1
    ws_s.row_dimensions[r].height = 22
    mcell(ws_s, r, 1, max(TB_LAST, 1 + len(all_per)),
          'PORTFOLIO MASTER DASHBOARD',
          F(bold=True, color=th['header_text'], size=13),
          Fill(th['header']), Aln('center'))
    r += 2

    # ── TABLE B ───────────────────────────────────────────────────────────────
    ws_s.row_dimensions[r].height = 18
    mcell(ws_s, r, 1, TB_LAST,
          'TABLE B — PROPERTY ASSET INFORMATION & YIELD ANALYSIS',
          F(bold=True, color=th['header_text'], size=10),
          Fill(th['header']), Aln('center'))
    r += 1

    ws_s.row_dimensions[r].height = 36
    for ci, h in enumerate(TB_HEADERS, 1):
        wcell(ws_s, r, ci, h,
              F(bold=True, color=th['header_text'], size=9),
              Fill(th['section']), Aln('center', wrap=True))
    r += 1

    TB_DATA_START = r
    prop_summary_rows = {}

    noi_pl  = LABEL_ROW.get('NOI (Net Operating Income)')
    ti_pl   = LABEL_ROW.get('Total Income')
    np_pl   = LABEL_ROW.get('NET PROFIT / (LOSS)')
    fin_pl  = LABEL_ROW.get('Total Financing Cost')
    prin_pl = LABEL_ROW.get('Principal Repaid')

    latest_fy = fy_labels[-1]
    c_probe   = 2
    for fy in fy_labels:
        if fy == latest_fy:
            latest_fy_col = c_probe
            break
        c_probe += 1 + 12

    for tab in prop_tabs:
        prop_summary_rows[tab] = r
        ws_s.row_dimensions[r].height = 18

        pinfo        = purchase_info.get(tab, {})
        prop_display = next((p['name'] for p in properties if p.get('tab') == tab), tab)

        wcell(ws_s, r, 1, prop_display, F(bold=True, size=9), aln=Aln('left'))
        wcell(ws_s, r, 2, pinfo.get('address', ''),
              F(color=INPUT_BLUE, size=9, italic=True), aln=Aln('left'))
        wcell(ws_s, r, 3, pinfo.get('purchase_price'),
              F(color=INPUT_BLUE, size=9), aln=Aln('right'), num_fmt=NUM_INT)
        wcell(ws_s, r, 4, pinfo.get('purchase_date'),
              F(color=INPUT_BLUE, size=9), aln=Aln('center'))
        wcell(ws_s, r, 5, pinfo.get('current_value'),
              F(color=INPUT_BLUE, size=9), aln=Aln('right'), num_fmt=NUM_INT)
        wcell(ws_s, r, 6, pinfo.get('mortgage'),
              F(color=INPUT_BLUE, size=9), aln=Aln('right'), num_fmt=NUM_INT)

        row_ref = str(r)
        wcell(ws_s, r, 7, f'=IFERROR(E{row_ref}-F{row_ref},"-")',
              F(size=9), aln=Aln('right'), num_fmt=NUM_INT)
        wcell(ws_s, r, 8, f'=IFERROR(F{row_ref}/E{row_ref},"-")',
              F(size=9), aln=Aln('right'), num_fmt=NUM_PERCENT)

        lc    = col(latest_fy_col)
        tab_q = f"'{tab}'" if '#' in tab or ' ' in tab else tab
        if ti_pl:
            wcell(ws_s, r, 9,
                  f'=IFERROR({tab_q}!{lc}{ti_pl}/C{row_ref},"-")',
                  F(size=9, color=GREEN_LINK), aln=Aln('right'), num_fmt=NUM_PERCENT)
        if np_pl:
            wcell(ws_s, r, 10,
                  f'=IFERROR({tab_q}!{lc}{np_pl}/C{row_ref},"-")',
                  F(size=9, color=GREEN_LINK), aln=Aln('right'), num_fmt=NUM_PERCENT)
        if noi_pl and fin_pl:
            dscr_f = (f'=IFERROR({tab_q}!{lc}{noi_pl}/'
                      f'({tab_q}!{lc}{fin_pl}+{tab_q}!{lc}{prin_pl}),"-")'
                      if prin_pl else
                      f'=IFERROR({tab_q}!{lc}{noi_pl}/{tab_q}!{lc}{fin_pl},"-")')
            wcell(ws_s, r, 11, dscr_f,
                  F(size=9, color=GREEN_LINK), aln=Aln('right'), num_fmt=NUM_DSCR)
        r += 1

    # Portfolio Total row (Table B)
    ws_s.row_dimensions[r].height = 18
    wcell(ws_s, r, 1, 'Portfolio Total',
          F(bold=True, color=th['header_text'], size=9),
          Fill(th['section']), Aln('left'))
    for ci in range(2, TB_LAST + 1):
        cl = col(ci)
        if ci in (2, 4):
            wcell(ws_s, r, ci, None, fill=Fill(th['section']))
            continue
        if ci in (3, 5, 6, 7):
            formula = f'=SUM({cl}{TB_DATA_START}:{cl}{r-1})'
            nm = NUM_INT
        elif ci == 8:
            formula = f'=IFERROR(SUM(F{TB_DATA_START}:F{r-1})/SUM(E{TB_DATA_START}:E{r-1}),"-")'
            nm = NUM_PERCENT
        elif ci in (9, 10):
            ti_parts = '+'.join([f"'{t}'!{col(latest_fy_col)}{ti_pl}" for t in prop_tabs]) if ti_pl else '0'
            np_parts = '+'.join([f"'{t}'!{col(latest_fy_col)}{np_pl}" for t in prop_tabs]) if np_pl else '0'
            numerator = ti_parts if ci == 9 else np_parts
            formula   = f'=IFERROR(({numerator})/SUM(C{TB_DATA_START}:C{r-1}),"-")'
            nm        = NUM_PERCENT
        elif ci == 11:
            noi_p = '+'.join([f"'{t}'!{col(latest_fy_col)}{noi_pl}" for t in prop_tabs]) if noi_pl else '0'
            fin_p = '+'.join([f"'{t}'!{col(latest_fy_col)}{fin_pl}" for t in prop_tabs]) if fin_pl else '0'
            pri_p = '+'.join([f"'{t}'!{col(latest_fy_col)}{prin_pl}" for t in prop_tabs]) if prin_pl else '0'
            formula   = f'=IFERROR(({noi_p})/(({fin_p})+({pri_p})),"-")'
            nm        = NUM_DSCR
        else:
            formula, nm = None, None

        wcell(ws_s, r, ci, formula,
              F(bold=True, color=th['header_text'], size=9),
              Fill(th['section']), Aln('right'), nm)
    r += 2

    # ── TABLE A ───────────────────────────────────────────────────────────────
    TA_LAST = 1 + len(all_per)

    ws_s.row_dimensions[r].height = 18
    mcell(ws_s, r, 1, TA_LAST,
          'TABLE A — PORTFOLIO PERFORMANCE SUMMARY (Linked from Property Tabs)',
          F(bold=True, color=th['header_text'], size=10),
          Fill(th['header']), Aln('center'))
    r += 1

    ws_s.row_dimensions[r].height = 36
    wcell(ws_s, r, 1, 'Metric / Property',
          F(bold=True, color=th['header_text'], size=9),
          Fill(th['section']), Aln('center', wrap=True))

    period_tab_col = {}
    c_probe = 2
    for fy in fy_labels:
        period_tab_col[f'FY {fy}'] = c_probe
        c_probe += 1 + 12
    for cy in cy_labels:
        period_tab_col[f'CY {cy}'] = c_probe
        c_probe += 1

    for ci, pk in enumerate(all_per, 2):
        is_cy = pk.startswith('CY')
        bg    = CY_YELLOW if is_cy else FY_YELLOW
        wcell(ws_s, r, ci, pk, F(bold=True, size=9), Fill(bg),
              Aln('center', wrap=True))
    r += 1

    ta_kpis = [
        ('Gross Rental Income', ti_pl,  NUM_INT),
        ('NOI',                 noi_pl, NUM_INT),
        ('NOI Margin %',        LABEL_ROW.get('NOI Margin %'), NUM_PERCENT),
        ('Financing Cost',      fin_pl, NUM_INT),
        ('Net Profit / (Loss)', np_pl,  NUM_INT),
        ('Net Cash Flow',       LABEL_ROW.get('Net Cash Flow'), NUM_INT),
    ]

    for kpi_disp, pl_row_num, nm in ta_kpis:
        ws_s.row_dimensions[r].height = 16
        mcell(ws_s, r, 1, TA_LAST, kpi_disp.upper(),
              F(bold=True, color=th['header_text'], size=9),
              Fill(th['section']), Aln('left'))
        r += 1

        prop_start_r = r
        for tab in prop_tabs:
            ws_s.row_dimensions[r].height = 14
            prop_display = next((p['name'] for p in properties if p.get('tab') == tab), tab)
            wcell(ws_s, r, 1, f'  {prop_display}', F(size=9), aln=Aln('left'))
            tab_q = f"'{tab}'" if '#' in tab or ' ' in tab else tab
            for ci, pk in enumerate(all_per, 2):
                tc    = period_tab_col.get(pk)
                is_cy = pk.startswith('CY')
                bg    = CY_YELLOW if is_cy else FY_YELLOW
                formula = f"={tab_q}!{col(tc)}{pl_row_num}" if tc and pl_row_num else None
                wcell(ws_s, r, ci, formula,
                      F(size=9, color=GREEN_LINK), Fill(bg), Aln('right'), nm)
            r += 1

        ws_s.row_dimensions[r].height = 15
        wcell(ws_s, r, 1, '  Portfolio Total',
              F(bold=True, size=9), Fill(th['kpi_row_bg']), Aln('left'))
        for ci, pk in enumerate(all_per, 2):
            cl    = col(ci)
            is_cy = pk.startswith('CY')
            bg    = CY_YELLOW if is_cy else FY_YELLOW
            formula = (f'=IFERROR(AVERAGE({cl}{prop_start_r}:{cl}{r-1}),"-")'
                       if kpi_disp == 'NOI Margin %' else
                       f'=SUM({cl}{prop_start_r}:{cl}{r-1})')
            wcell(ws_s, r, ci, formula,
                  F(bold=True, size=9), Fill(bg), Aln('right'), nm)
        r += 1

    # ── Summary column widths ─────────────────────────────────────────────────
    ws_s.column_dimensions['A'].width = 30
    ws_s.column_dimensions['B'].width = 22
    ws_s.column_dimensions['C'].width = 16
    ws_s.column_dimensions['D'].width = 13
    ws_s.column_dimensions['E'].width = 16
    ws_s.column_dimensions['F'].width = 16
    ws_s.column_dimensions['G'].width = 13
    ws_s.column_dimensions['H'].width = 10
    ws_s.column_dimensions['I'].width = 13
    ws_s.column_dimensions['J'].width = 13
    ws_s.column_dimensions['K'].width = 12
    for ci in range(2, 2 + len(all_per)):
        ws_s.column_dimensions[col(ci)].width = 13

    ws_s.freeze_panes = 'B4'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
